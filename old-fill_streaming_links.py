#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fill_streaming_links.py
========================

Complète automatiquement l'onglet "Liens streaming" de DJMashupLibrary.xlsx
(colonnes YouTube clip officiel / YouTube audio officiel / Spotify / Deezer)
en interrogeant les APIs officielles de chaque plateforme. Aucune IA n'est
utilisée ici : uniquement des appels HTTP directs + une logique de scoring
simple pour choisir le meilleur résultat.

PRÉREQUIS
---------
1. Python 3.8+ avec les paquets :
       pip install requests openpyxl

2. Des clés API (gratuites) pour YouTube et Spotify :

   - YouTube Data API v3 :
       a) https://console.cloud.google.com/ -> créer un projet
       b) Activer "YouTube Data API v3"
       c) Créer une clé API (APIs & Services > Identifiants)
       d) Quota gratuit : 10 000 unités/jour, une recherche = 100 unités
          -> environ 100 recherches/jour avec le quota gratuit par défaut.
             Pour 6370 titres, prévoir plusieurs jours OU demander une
             augmentation de quota à Google (formulaire gratuit).

   - Spotify Web API :
       a) https://developer.spotify.com/dashboard -> créer une app
       b) Récupérer Client ID + Client Secret
       c) Pas de quota journalier strict, juste du rate-limiting normal.

   - Deezer : aucune clé nécessaire (API publique en lecture).

3. Renseigner les clés soit en variables d'environnement, soit directement
   dans la section CONFIG ci-dessous :

       set YOUTUBE_API_KEY=xxxx        (Windows cmd)
       set SPOTIFY_CLIENT_ID=xxxx
       set SPOTIFY_CLIENT_SECRET=xxxx

       $env:YOUTUBE_API_KEY="xxxx"     (PowerShell)

UTILISATION
-----------
    python fill_streaming_links.py

Le script :
  - lit DJMashupLibrary.xlsx (onglet "Liens streaming"), colonne ID/Artiste/Titre
  - saute les lignes déjà complètes (les 3 liens déjà renseignés)
  - interroge Deezer (toujours), Spotify (si clés fournies), YouTube (si clé fournie)
  - écrit les résultats directement dans les colonnes correspondantes
  - sauvegarde le fichier tous les SAVE_EVERY morceaux traités (reprise possible
    si le script est interrompu : il repart du premier trou trouvé)
  - journalise sa progression dans la console

Le script est volontairement PRUDENT sur le rythme des requêtes (throttling)
pour ne pas se faire bannir temporairement par les APIs.
"""

import os
import sys
import time
import json
import re
import requests
import openpyxl

# ───────────────────────────── CONFIG ──────────────────────────────────────

XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "DJMashupLibrary.xlsx")
SHEET_NAME = "Liens streaming"

YOUTUBE_API_KEY = os.environ.get("YOUTUBE_API_KEY", "")
SPOTIFY_CLIENT_ID = os.environ.get("SPOTIFY_CLIENT_ID", "")
SPOTIFY_CLIENT_SECRET = os.environ.get("SPOTIFY_CLIENT_SECRET", "")

# Sauvegarde le classeur tous les N morceaux traités (protège contre une
# interruption en cours de route — quota YouTube épuisé, coupure réseau...).
SAVE_EVERY = 15

# Pause entre 2 morceaux (secondes) — reste raisonnable vis-à-vis des 3 APIs.
SLEEP_BETWEEN_TRACKS = 0.3

# Colonnes de l'onglet "Liens streaming" (1-indexé, cf. en-têtes du fichier) :
#   A=ID  B=Artiste  C=Titre  D=YouTube clip officiel  E=YouTube audio officiel
#   F=Spotify  G=Deezer  H=Apple Music  I=Beatport  J=Discogs  K=WhoSampled  L=Genius
COL_ID, COL_ARTIST, COL_TITLE = 1, 2, 3
COL_YT_CLIP, COL_YT_AUDIO, COL_SPOTIFY, COL_DEEZER = 4, 5, 6, 7

# ─────────────────────────── UTILITAIRES ───────────────────────────────────

def log(msg):
    print(f"[fill_streaming_links] {msg}", flush=True)


def clean(s):
    return (s or "").strip()


# ─────────────────────────── DEEZER (pas de clé) ───────────────────────────

def search_deezer(artist, title):
    """Retourne le lien deezer.com du meilleur résultat, ou None."""
    try:
        r = requests.get(
            "https://api.deezer.com/search",
            params={"q": f'artist:"{artist}" track:"{title}"'},
            timeout=10,
        )
        r.raise_for_status()
        data = r.json()
        results = data.get("data", [])
        if not results:
            # repli : recherche moins stricte (sans les guillemets exacts)
            r = requests.get(
                "https://api.deezer.com/search",
                params={"q": f"{artist} {title}"},
                timeout=10,
            )
            r.raise_for_status()
            results = r.json().get("data", [])
        if results:
            return results[0].get("link")
    except requests.RequestException as e:
        log(f"  ⚠ Deezer erreur : {e}")
    return None


# ─────────────────────────── SPOTIFY (Client Credentials) ─────────────────

_spotify_token = None
_spotify_token_expiry = 0


def get_spotify_token():
    global _spotify_token, _spotify_token_expiry
    if not SPOTIFY_CLIENT_ID or not SPOTIFY_CLIENT_SECRET:
        return None
    if _spotify_token and time.time() < _spotify_token_expiry:
        return _spotify_token
    try:
        r = requests.post(
            "https://accounts.spotify.com/api/token",
            data={"grant_type": "client_credentials"},
            auth=(SPOTIFY_CLIENT_ID, SPOTIFY_CLIENT_SECRET),
            timeout=10,
        )
        r.raise_for_status()
        payload = r.json()
        _spotify_token = payload["access_token"]
        _spotify_token_expiry = time.time() + payload.get("expires_in", 3600) - 60
        return _spotify_token
    except requests.RequestException as e:
        log(f"  ⚠ Spotify auth erreur : {e}")
        return None


def search_spotify(artist, title):
    """Retourne le lien open.spotify.com du meilleur résultat, ou None."""
    token = get_spotify_token()
    if not token:
        return None
    try:
        r = requests.get(
            "https://api.spotify.com/v1/search",
            headers={"Authorization": f"Bearer {token}"},
            params={"q": f"track:{title} artist:{artist}", "type": "track", "limit": 5},
            timeout=10,
        )
        if r.status_code == 401:
            # token expiré entre-temps -> on force un renouvellement puis on retente une fois
            global _spotify_token
            _spotify_token = None
            token = get_spotify_token()
            if not token:
                return None
            r = requests.get(
                "https://api.spotify.com/v1/search",
                headers={"Authorization": f"Bearer {token}"},
                params={"q": f"track:{title} artist:{artist}", "type": "track", "limit": 5},
                timeout=10,
            )
        r.raise_for_status()
        items = r.json().get("tracks", {}).get("items", [])
        if items:
            return items[0]["external_urls"]["spotify"]
    except requests.RequestException as e:
        log(f"  ⚠ Spotify erreur : {e}")
    return None


# ─────────────────────────── YOUTUBE (Data API v3) ─────────────────────────

# Mots-clés qui trahissent un clip "officiel" dans le titre de la vidéo ou le
# nom de la chaîne — sert à écarter les reprises fan-made / lyrics videos de
# qualité douteuse, sans dépendre d'une IA.
OFFICIAL_HINTS = ("official video", "official music video", "official audio", "clip officiel")
BAD_HINTS = ("cover", "reprise", "karaoke", "tribute", "instrumental", "lyrics only", "8d audio", "nightcore", "sped up", "slowed")


def score_youtube_item(item, artist):
    title = item["snippet"]["title"].lower()
    channel = item["snippet"]["channelTitle"].lower()
    score = 0
    if any(h in title for h in OFFICIAL_HINTS):
        score += 10
    if any(b in title for b in BAD_HINTS):
        score -= 10
    # Chaîne qui porte le nom de l'artiste (ou VEVO, très souvent la source
    # officielle) -> bon signal de légitimité.
    artist_lower = artist.lower()
    if artist_lower in channel or "vevo" in channel:
        score += 5
    return score


def search_youtube(artist, title):
    """Retourne le lien youtube.com/watch?v=... du meilleur résultat, ou None."""
    if not YOUTUBE_API_KEY:
        return None
    try:
        r = requests.get(
            "https://www.googleapis.com/youtube/v3/search",
            params={
                "part": "snippet",
                "q": f"{artist} {title} official video",
                "type": "video",
                "maxResults": 8,
                "key": YOUTUBE_API_KEY,
            },
            timeout=10,
        )
        r.raise_for_status()
        items = r.json().get("items", [])
        if not items:
            return None
        best = max(items, key=lambda it: score_youtube_item(it, artist))
        video_id = best["id"]["videoId"]
        return f"https://www.youtube.com/watch?v={video_id}"
    except requests.RequestException as e:
        log(f"  ⚠ YouTube erreur : {e}")
        return None


# ─────────────────────────────── MAIN ──────────────────────────────────────

def main():
    if not os.path.exists(XLSX_PATH):
        log(f"❌ Fichier introuvable : {XLSX_PATH}")
        sys.exit(1)

    if not YOUTUBE_API_KEY:
        log("⚠ YOUTUBE_API_KEY absente -> les liens YouTube seront laissés vides.")
    if not SPOTIFY_CLIENT_ID or not SPOTIFY_CLIENT_SECRET:
        log("⚠ SPOTIFY_CLIENT_ID/SECRET absents -> les liens Spotify seront laissés vides.")

    log(f"Ouverture de {XLSX_PATH} ...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    total_rows = ws.max_row
    processed = 0
    skipped = 0

    for row_idx in range(2, total_rows + 1):
        row = ws[row_idx]
        track_id = row[COL_ID - 1].value
        artist = clean(row[COL_ARTIST - 1].value)
        title = clean(row[COL_TITLE - 1].value)

        if not artist or not title:
            continue  # ligne vide (fin de tableau)

        already_yt = clean(row[COL_YT_CLIP - 1].value)
        already_sp = clean(row[COL_SPOTIFY - 1].value)
        already_dz = clean(row[COL_DEEZER - 1].value)

        # Ligne déjà complète (les 3 liens principaux présents) -> on saute,
        # ce qui rend le script relançable sans dupliquer le travail déjà fait.
        if already_yt and already_sp and already_dz:
            skipped += 1
            continue

        log(f"[{row_idx - 1}/{total_rows - 1}] {artist} - {title}")

        if not already_dz:
            dz = search_deezer(artist, title)
            if dz:
                row[COL_DEEZER - 1].value = dz
                log(f"  ✓ Deezer  : {dz}")
            else:
                log("  ✗ Deezer  : rien trouvé")

        if not already_sp:
            sp = search_spotify(artist, title)
            if sp:
                row[COL_SPOTIFY - 1].value = sp
                log(f"  ✓ Spotify : {sp}")
            elif SPOTIFY_CLIENT_ID:
                log("  ✗ Spotify : rien trouvé")

        if not already_yt:
            yt = search_youtube(artist, title)
            if yt:
                row[COL_YT_CLIP - 1].value = yt
                log(f"  ✓ YouTube : {yt}")
            elif YOUTUBE_API_KEY:
                log("  ✗ YouTube : rien trouvé")

        processed += 1
        time.sleep(SLEEP_BETWEEN_TRACKS)

        if processed % SAVE_EVERY == 0:
            wb.save(XLSX_PATH)
            log(f"💾 Sauvegarde intermédiaire ({processed} morceaux traités dans cette session)")

    wb.save(XLSX_PATH)
    log(f"✅ Terminé. {processed} morceaux traités, {skipped} déjà complets ignorés.")


if __name__ == "__main__":
    main()

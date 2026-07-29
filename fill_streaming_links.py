#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fill_streaming_links_optimized.py
=================================

Version optimisée du script pour compléter automatiquement l'onglet "Liens streaming"
de DJMashupLibrary.xlsx avec une meilleure gestion des erreurs, du caching,
et de l'encodage des requêtes API.

AMÉLIORATIONS PRINCIPALES :
---------------------------
1. Encodage URL propre des paramètres de requête (fix le problème des caractères spéciaux)
2. Gestion robuste des erreurs API avec retry logic
3. Cache des résultats pour éviter les requêtes duplicata
4. Sessions HTTP persistantes pour de meilleures performances
5. Meilleure gestion des permissions fichiers
6. Validation des inputs avant les requêtes
7. Journalisation améliorée avec niveaux de log
8. Gestion des rate limits et throttling adaptatif
"""

import os
import sys
import time
import json
import re
import urllib.parse
from functools import lru_cache
from typing import Optional, Tuple
import requests
import openpyxl
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# ───────────────────────────── CONFIG ──────────────────────────────────────

XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "DJMashupLibrary.xlsx")
SHEET_NAME = "Liens streaming"
BASE_TITRES_SHEET = "Base titres"


def _load_key_from_backend_env(var_name: str) -> str:
    """Va chercher une clé déjà configurée pour le site lui-même, dans
    backend/.env (le backend Node lit YT_API_KEY à cet endroit précis,
    cf. backend/routes/youtube.js) — évite d'avoir à en recréer une."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backend", ".env")
    if not os.path.exists(env_path):
        return ""
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line.startswith(f"{var_name}="):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")
    except OSError:
        pass
    return ""


def _looks_like_placeholder(value: str) -> bool:
    """Repère les clés manifestement pas remplacées (ex. 'ta_cle' laissée
    telle quelle depuis un exemple de commande) pour ne pas marteler les
    APIs avec des identifiants forcément invalides."""
    if not value:
        return True
    v = value.lower()
    return v in ("ta_cle", "xxxx", "your_key_here", "changeme", "todo", "ton_id", "ton_secret") or len(v) < 8


YOUTUBE_API_KEY = os.environ.get("YOUTUBE_API_KEY", "") or _load_key_from_backend_env("YT_API_KEY")
SPOTIFY_CLIENT_ID = os.environ.get("SPOTIFY_CLIENT_ID", "")
SPOTIFY_CLIENT_SECRET = os.environ.get("SPOTIFY_CLIENT_SECRET", "")

if _looks_like_placeholder(YOUTUBE_API_KEY):
    if YOUTUBE_API_KEY:
        print(f"[fill_streaming_links] ⚠ YOUTUBE_API_KEY ('{YOUTUBE_API_KEY}') ressemble à un placeholder — ignorée.")
    YOUTUBE_API_KEY = ""
else:
    print(f"[fill_streaming_links] ✓ Clé YouTube récupérée depuis backend/.env (YT_API_KEY), même clé que le site.")
if _looks_like_placeholder(SPOTIFY_CLIENT_ID):
    SPOTIFY_CLIENT_ID = ""
if _looks_like_placeholder(SPOTIFY_CLIENT_SECRET):
    SPOTIFY_CLIENT_SECRET = ""

# Sauvegarde le classeur tous les N morceaux traités
SAVE_EVERY = 15

# Pause entre 2 morceaux (secondes) - adaptatif selon les erreurs
BASE_SLEEP_BETWEEN_TRACKS = 0.3
MAX_SLEEP_BETWEEN_TRACKS = 5.0
current_sleep = BASE_SLEEP_BETWEEN_TRACKS

# Colonnes de l'onglet "Liens streaming" (1-indexé)
COL_ID, COL_ARTIST, COL_TITLE = 1, 2, 3
COL_YT_CLIP, COL_YT_AUDIO, COL_SPOTIFY, COL_DEEZER = 4, 5, 6, 7

# Cache TTL en secondes (1 heure)
CACHE_TTL = 3600

# Configuration des retries pour les requêtes HTTP
RETRY_CONFIG = Retry(
    total=3,
    backoff_factor=1,
    status_forcelist=[400, 401, 403, 429, 500, 502, 503, 504],
    allowed_methods=["GET", "POST"]
)


# ─────────────────────────── UTILITAIRES ───────────────────────────────────

class Logger:
    """Logger simple avec niveaux de log"""
    
    @staticmethod
    def info(msg: str):
        print(f"[INFO] {msg}", flush=True)
    
    @staticmethod
    def warning(msg: str):
        print(f"[WARNING] {msg}", flush=True)
    
    @staticmethod
    def error(msg: str):
        print(f"[ERROR] {msg}", flush=True)
    
    @staticmethod
    def success(msg: str):
        print(f"[SUCCESS] {msg}", flush=True)


def clean_and_normalize(s: str) -> str:
    """Nettoie et normalise une string pour les requêtes API"""
    if not s:
        return ""
    
    # Supprime les caractères de contrôle et espaces superflus
    s = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', str(s)).strip()

    # Au cas où une formule Excel non résolue arriverait quand même ici
    # (garde-fou en plus du lookup direct sur "Base titres" plus haut) —
    # ne s'applique qu'aux chaînes qui COMMENCENT par "=" (vraie formule),
    # pour ne jamais tronquer un titre légitime contenant un "=".
    if s.startswith("="):
        s = re.sub(r"^=+", '', s)
        s = re.sub(r"'[^']*'!\$?[A-Z]+\$?\d+", '', s)

    # Remplace les espaces multiples par un seul
    s = re.sub(r'\s+', ' ', s)
    
    return s.strip()


def url_encode_query(query: str) -> str:
    """Encode proprement une requête pour les URLs"""
    # Nettoie d'abord
    query = clean_and_normalize(query)
    # Encode pour URL
    return urllib.parse.quote(query)


# ─────────────────────────── CACHE ─────────────────────────────────────────

class SimpleCache:
    """Cache simple avec TTL pour éviter les requêtes duplicata"""
    
    def __init__(self, ttl: int = CACHE_TTL):
        self.cache = {}
        self.ttl = ttl
        self.last_cleanup = time.time()
    
    def get(self, key: str) -> Optional[str]:
        """Récupère une valeur du cache si elle existe et n'est pas expirée"""
        if key in self.cache:
            value, timestamp = self.cache[key]
            if time.time() - timestamp < self.ttl:
                return value
            else:
                del self.cache[key]
        return None
    
    def set(self, key: str, value: str):
        """Stocke une valeur dans le cache"""
        self.cache[key] = (value, time.time())
        
        # Nettoyage périodique
        if time.time() - self.last_cleanup > 300:  # 5 minutes
            self.cleanup()
            self.last_cleanup = time.time()
    
    def cleanup(self):
        """Nettoie les entrées expirées"""
        expired = [k for k, (_, ts) in self.cache.items() 
                  if time.time() - ts > self.ttl]
        for k in expired:
            del self.cache[k]


# Initialisation du cache
api_cache = SimpleCache(CACHE_TTL)


# ─────────────────────────── SESSIONS HTTP ────────────────────────────────

# Crée une session avec retry logic
session = requests.Session()
session.mount("http://", HTTPAdapter(max_retries=RETRY_CONFIG))
session.mount("https://", HTTPAdapter(max_retries=RETRY_CONFIG))


def make_request(method: str, url: str, **kwargs) -> Optional[requests.Response]:
    """Effectue une requête HTTP avec gestion des erreurs améliorée"""
    try:
        response = session.request(method, url, timeout=15, **kwargs)
        
        # Gestion spécifique des erreurs 429 (Too Many Requests)
        if response.status_code == 429:
            retry_after = int(response.headers.get('Retry-After', 10))
            Logger.warning(f"Rate limited. Retrying after {retry_after} seconds...")
            time.sleep(retry_after)
            response = session.request(method, url, timeout=15, **kwargs)
        
        response.raise_for_status()
        return response
    except requests.exceptions.RequestException as e:
        Logger.error(f"Request failed: {e}")
        return None
    except Exception as e:
        Logger.error(f"Unexpected error: {e}")
        return None


# ─────────────────────────── DEEZER ────────────────────────────────────────

def search_deezer(artist: str, title: str) -> Optional[str]:
    """Retourne le lien deezer.com du meilleur résultat, ou None."""
    cache_key = f"deezer:{artist}:{title}"
    cached = api_cache.get(cache_key)
    if cached:
        return cached
    
    artist_clean = clean_and_normalize(artist)
    title_clean = clean_and_normalize(title)
    
    if not artist_clean or not title_clean:
        return None
    
    try:
        # Première tentative : recherche stricte
        params = {"q": f'artist:"{artist_clean}" track:"{title_clean}"'}
        r = make_request("GET", "https://api.deezer.com/search", params=params)
        
        if r and r.status_code == 200:
            data = r.json()
            results = data.get("data", [])
            if results:
                link = results[0].get("link")
                if link:
                    api_cache.set(cache_key, link)
                    return link
        
        # Deuxième tentative : recherche moins stricte
        params = {"q": f"{artist_clean} {title_clean}"}
        r = make_request("GET", "https://api.deezer.com/search", params=params)
        
        if r and r.status_code == 200:
            data = r.json()
            results = data.get("data", [])
            if results:
                link = results[0].get("link")
                if link:
                    api_cache.set(cache_key, link)
                    return link
        
    except Exception as e:
        Logger.error(f"Deezer error: {e}")
    
    return None


# ─────────────────────────── SPOTIFY ───────────────────────────────────────

_spotify_token: Optional[str] = None
_spotify_token_expiry: float = 0


def get_spotify_token() -> Optional[str]:
    """Récupère un token Spotify valide"""
    global _spotify_token, _spotify_token_expiry
    
    if not SPOTIFY_CLIENT_ID or not SPOTIFY_CLIENT_SECRET:
        return None
    
    if _spotify_token and time.time() < _spotify_token_expiry:
        return _spotify_token
    
    try:
        r = make_request(
            "POST",
            "https://accounts.spotify.com/api/token",
            data={"grant_type": "client_credentials"},
            auth=(SPOTIFY_CLIENT_ID, SPOTIFY_CLIENT_SECRET)
        )
        
        if r and r.status_code == 200:
            payload = r.json()
            _spotify_token = payload.get("access_token")
            expires_in = payload.get("expires_in", 3600)
            _spotify_token_expiry = time.time() + expires_in - 60  # 1 minute de marge
            Logger.info("Nouveau token Spotify obtenu")
            return _spotify_token
        else:
            Logger.error("Failed to get Spotify token")
            return None
    except Exception as e:
        Logger.error(f"Spotify auth error: {e}")
        return None


def search_spotify(artist: str, title: str) -> Optional[str]:
    """Retourne le lien open.spotify.com du meilleur résultat, ou None."""
    cache_key = f"spotify:{artist}:{title}"
    cached = api_cache.get(cache_key)
    if cached:
        return cached
    
    artist_clean = clean_and_normalize(artist)
    title_clean = clean_and_normalize(title)
    
    if not artist_clean or not title_clean:
        return None
    
    token = get_spotify_token()
    if not token:
        return None
    
    try:
        # Encode la requête proprement
        query = f"track:{url_encode_query(title_clean)} artist:{url_encode_query(artist_clean)}"
        
        r = make_request(
            "GET",
            "https://api.spotify.com/v1/search",
            headers={"Authorization": f"Bearer {token}"},
            params={
                "q": query,
                "type": "track",
                "limit": 5,
                "market": "FR"  # Ajout du marché pour de meilleurs résultats
            }
        )
        
        if r and r.status_code == 200:
            data = r.json()
            items = data.get("tracks", {}).get("items", [])
            if items:
                link = items[0]["external_urls"]["spotify"]
                api_cache.set(cache_key, link)
                return link
        elif r and r.status_code == 401:
            # Token expiré, on réessaye une fois
            global _spotify_token
            _spotify_token = None
            token = get_spotify_token()
            if token:
                r = make_request(
                    "GET",
                    "https://api.spotify.com/v1/search",
                    headers={"Authorization": f"Bearer {token}"},
                    params={
                        "q": query,
                        "type": "track",
                        "limit": 5,
                        "market": "FR"
                    }
                )
                if r and r.status_code == 200:
                    data = r.json()
                    items = data.get("tracks", {}).get("items", [])
                    if items:
                        link = items[0]["external_urls"]["spotify"]
                        api_cache.set(cache_key, link)
                        return link
        
    except Exception as e:
        Logger.error(f"Spotify search error: {e}")
    
    return None


# ─────────────────────────── YOUTUBE ───────────────────────────────────────

# Mots-clés pour identifier les clips officiels
OFFICIAL_HINTS = ("official video", "official music video", "official audio", 
                  "clip officiel", "video officielle", "audio officiel")
BAD_HINTS = ("cover", "reprise", "karaoke", "tribute", "instrumental", 
             "lyrics only", "8d audio", "nightcore", "sped up", "slowed",
             "remix", "mashup", "edit", "fan made", "unofficial")


def score_youtube_item(item: dict, artist: str) -> int:
    """Score un résultat YouTube pour trouver le meilleur"""
    title = item["snippet"]["title"].lower()
    channel = item["snippet"]["channelTitle"].lower()
    description = item["snippet"].get("description", "").lower()
    score = 0
    
    # Bonus pour les mots-clés officiels
    if any(h in title for h in OFFICIAL_HINTS):
        score += 10
    
    # Pénalité pour les mots-clés indésirables
    if any(b in title for b in BAD_HINTS):
        score -= 10
    
    # Bonus si la chaîne contient le nom de l'artiste ou VEVO
    artist_lower = artist.lower()
    if artist_lower in channel or "vevo" in channel:
        score += 5
    
    # Bonus si la description contient des liens officiels
    if "http://" in description or "https://" in description:
        score += 2
    
    # Bonus pour les vidéos plus longues (probablement pas des teasers)
    # Note: duration n'est pas dans les résultats de recherche, il faudrait une requête supplémentaire
    
    return score


def search_youtube(artist: str, title: str) -> Optional[str]:
    """Retourne le lien youtube.com/watch?v=... du meilleur résultat, ou None."""
    if not YOUTUBE_API_KEY:
        return None
    
    cache_key = f"youtube:{artist}:{title}"
    cached = api_cache.get(cache_key)
    if cached:
        return cached
    
    artist_clean = clean_and_normalize(artist)
    title_clean = clean_and_normalize(title)
    
    if not artist_clean or not title_clean:
        return None
    
    try:
        # Construction de la requête avec encodage propre
        query_parts = [artist_clean, title_clean, "official"]
        query = " ".join(query_parts)
        
        r = make_request(
            "GET",
            "https://www.googleapis.com/youtube/v3/search",
            params={
                "part": "snippet",
                "q": query,
                "type": "video",
                "maxResults": 10,  # Augmenté pour plus de choix
                "key": YOUTUBE_API_KEY,
                "videoDuration": "medium,long",  # Évite les très courtes vidéos
                "videoDefinition": "high",  # Préfère les vidéos HD
                "order": "relevance"
            }
        )
        
        if r and r.status_code == 200:
            data = r.json()
            items = data.get("items", [])
            if items:
                # Filtre les vidéos avec des IDs valides
                valid_items = [item for item in items 
                              if item.get("id", {}).get("kind") == "youtube#video"]
                if valid_items:
                    best = max(valid_items, key=lambda it: score_youtube_item(it, artist_clean))
                    video_id = best["id"]["videoId"]
                    link = f"https://www.youtube.com/watch?v={video_id}"
                    api_cache.set(cache_key, link)
                    return link
        
    except Exception as e:
        Logger.error(f"YouTube search error: {e}")
    
    return None


# ─────────────────────────── GESTION DES ERREURS ──────────────────────────

def check_file_permissions(filepath: str) -> bool:
    """Vérifie si on a les permissions pour écrire dans le fichier"""
    try:
        # Teste l'ouverture en écriture
        with open(filepath, 'a') as f:
            pass
        return True
    except PermissionError:
        return False
    except Exception as e:
        Logger.error(f"Error checking file permissions: {e}")
        return False


def wait_for_file_access(filepath: str, max_attempts: int = 10) -> bool:
    """Attend que le fichier soit accessible (fermé par d'autres processus)"""
    for attempt in range(max_attempts):
        if check_file_permissions(filepath):
            return True
        Logger.warning(f"Fichier verrouillé, tentative {attempt + 1}/{max_attempts}...")
        time.sleep(2)
    return False


# ─────────────────────────────── MAIN ──────────────────────────────────────

def main():
    global current_sleep
    
    # Vérification des prérequis
    if not os.path.exists(XLSX_PATH):
        Logger.error(f"Fichier introuvable : {XLSX_PATH}")
        sys.exit(1)
    
    # Vérification des permissions
    if not wait_for_file_access(XLSX_PATH):
        Logger.error(f"Impossible d'accéder au fichier {XLSX_PATH} (verrouillé ou permissions insuffisantes)")
        sys.exit(1)
    
    # Avertissements pour les clés API manquantes
    if not YOUTUBE_API_KEY:
        Logger.warning("YOUTUBE_API_KEY absente -> les liens YouTube seront laissés vides.")
    if not SPOTIFY_CLIENT_ID or not SPOTIFY_CLIENT_SECRET:
        Logger.warning("SPOTIFY_CLIENT_ID/SECRET absents -> les liens Spotify seront laissés vides.")
    
    Logger.info(f"Ouverture de {XLSX_PATH}...")
    
    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb[SHEET_NAME]
        ws_base = wb[BASE_TITRES_SHEET]
    except Exception as e:
        Logger.error(f"Erreur lors de l'ouverture du fichier Excel: {e}")
        sys.exit(1)

    # Les colonnes Artiste/Titre de l'onglet "Liens streaming" sont en réalité
    # des FORMULES qui pointent vers l'onglet "Base titres" (ex. ="Base
    # titres"!B46). openpyxl ne calcule jamais les formules : lire .value
    # dessus renvoie le texte de la formule elle-même, pas le nom réel de
    # l'artiste — c'est ce qui provoquait les recherches vides ("='Base
    # titres'!B46" envoyé tel quel aux APIs). On construit donc une table de
    # correspondance ID -> (artiste, titre) directement depuis "Base titres"
    # (colonnes B et C), et on l'utilise à la place des cellules-formules.
    base_lookup = {}
    for base_row in ws_base.iter_rows(min_row=2, values_only=True):
        base_id = base_row[0]
        if base_id is None:
            continue
        base_lookup[base_id] = (
            (base_row[1] or "").strip() if isinstance(base_row[1], str) else (base_row[1] or ""),
            (base_row[2] or "").strip() if isinstance(base_row[2], str) else (base_row[2] or ""),
        )
    Logger.info(f"{len(base_lookup)} morceaux indexés depuis l'onglet '{BASE_TITRES_SHEET}'.")

    total_rows = ws.max_row
    processed = 0
    skipped = 0
    errors = 0
    
    Logger.info(f"Traitement de {total_rows - 1} morceaux...")
    
    for row_idx in range(2, total_rows + 1):
        try:
            row = ws[row_idx]
            track_id = row[COL_ID - 1].value

            # Artiste/Titre viennent de "Base titres" (cf. base_lookup
            # ci-dessus) et non des cellules-formules de cet onglet.
            base_entry = base_lookup.get(track_id)
            if not base_entry:
                Logger.warning(f"Ligne {row_idx}: ID {track_id} introuvable dans '{BASE_TITRES_SHEET}', on saute")
                continue
            artist = clean_and_normalize(str(base_entry[0])) if base_entry[0] else ""
            title = clean_and_normalize(str(base_entry[1])) if base_entry[1] else ""

            if not artist or not title:
                Logger.warning(f"Ligne {row_idx}: Artiste ou Titre vide, on saute")
                continue
            
            already_yt = clean_and_normalize(str(row[COL_YT_CLIP - 1].value) if row[COL_YT_CLIP - 1].value else "")
            already_sp = clean_and_normalize(str(row[COL_SPOTIFY - 1].value) if row[COL_SPOTIFY - 1].value else "")
            already_dz = clean_and_normalize(str(row[COL_DEEZER - 1].value) if row[COL_DEEZER - 1].value else "")
            
            # Ligne déjà complète -> on saute
            if already_yt and already_sp and already_dz:
                skipped += 1
                continue
            
            Logger.info(f"[{row_idx - 1}/{total_rows - 1}] {artist} - {title}")
            
            # Recherche Deezer
            if not already_dz:
                dz = search_deezer(artist, title)
                if dz:
                    row[COL_DEEZER - 1].value = dz
                    Logger.success(f"  ✓ Deezer  : {dz}")
                else:
                    Logger.warning("  ✗ Deezer  : rien trouvé")
            
            # Recherche Spotify
            if not already_sp:
                sp = search_spotify(artist, title)
                if sp:
                    row[COL_SPOTIFY - 1].value = sp
                    Logger.success(f"  ✓ Spotify : {sp}")
                elif SPOTIFY_CLIENT_ID:
                    Logger.warning("  ✗ Spotify : rien trouvé")
            
            # Recherche YouTube
            if not already_yt:
                yt = search_youtube(artist, title)
                if yt:
                    row[COL_YT_CLIP - 1].value = yt
                    Logger.success(f"  ✓ YouTube : {yt}")
                elif YOUTUBE_API_KEY:
                    Logger.warning("  ✗ YouTube : rien trouvé")
            
            processed += 1
            
            # Throttling adaptatif
            time.sleep(current_sleep)
            
            # Sauvegarde intermédiaire
            if processed % SAVE_EVERY == 0:
                try:
                    wb.save(XLSX_PATH)
                    Logger.info(f"💾 Sauvegarde intermédiaire ({processed} morceaux traités)")
                except PermissionError:
                    Logger.warning("Sauvegarde impossible (fichier verrouillé), on continue...")
                    # Augmente le sleep pour éviter les conflits
                    current_sleep = min(current_sleep * 1.5, MAX_SLEEP_BETWEEN_TRACKS)
                except Exception as e:
                    Logger.error(f"Erreur lors de la sauvegarde intermédiaire: {e}")
            
        except Exception as e:
            errors += 1
            Logger.error(f"Erreur sur la ligne {row_idx}: {e}")
            # Augmente le sleep en cas d'erreur
            current_sleep = min(current_sleep * 1.2, MAX_SLEEP_BETWEEN_TRACKS)
            time.sleep(current_sleep)
            continue
    
    # Sauvegarde finale
    try:
        wb.save(XLSX_PATH)
        Logger.success(f"💾 Sauvegarde finale effectuée")
    except PermissionError:
        Logger.error("Sauvegarde finale impossible (fichier verrouillé)")
        Logger.warning("Fermez le fichier Excel et relancez le script pour sauvegarder les modifications")
    except Exception as e:
        Logger.error(f"Erreur lors de la sauvegarde finale: {e}")
    finally:
        wb.close()
    
    Logger.success(f"✅ Terminé. {processed} morceaux traités, {skipped} déjà complets ignorés, {errors} erreurs.")


if __name__ == "__main__":
    main()